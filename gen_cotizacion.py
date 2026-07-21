"""Genera CALCULADORA_PRECIOS_MINZOE.xlsx en el escritorio."""
import os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

ROJO   = "DC2626"
BLANCO = "FFFFFF"
GRIS   = "F9F9F9"
ROJO_L = "FFF0F0"

wb = Workbook()
ws = wb.active
ws.title = "CALCULADORA"

# ── Anchos de columna ─────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 20

# ── Helpers de estilo ─────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color=ROJO)
    return Border(left=s, right=s, top=s, bottom=s)

# ── ENCABEZADO EMPRESA ────────────────────────────────────────────────────
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value          = "CONSTRUCCIONES MINZOE SAS"
c.font           = Font(bold=True, size=14, color=BLANCO)
c.fill           = fill(ROJO)
c.alignment      = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:E2")
c = ws["A2"]
c.value     = "Cra 5 # 8a-18  |  3175102668 – 3173748665  |  construminzoe@gmail.com"
c.font      = Font(size=9, color=ROJO)
c.fill      = fill("FFF5F5")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

ws.merge_cells("A3:E3")
c = ws["A3"]
c.value     = "CALCULADORA DE PRECIOS"
c.font      = Font(bold=True, size=12, color=BLANCO)
c.fill      = fill(ROJO)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 22

ws.row_dimensions[4].height = 6  # separador

# ── ENCABEZADOS TABLA ─────────────────────────────────────────────────────
headers = ["#", "DESCRIPCIÓN DEL ÍTEM", "PRECIO PROVEEDOR", "% RENTABILIDAD", "PRECIO UNITARIO"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col)
    c.value     = h
    c.font      = Font(bold=True, size=10, color=BLANCO)
    c.fill      = fill(ROJO)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border_thin()
ws.row_dimensions[5].height = 28

# ── FILAS DE DATOS (20 ítems) ─────────────────────────────────────────────
FILAS = 20
for i in range(FILAS):
    r   = 6 + i
    bg  = GRIS if i % 2 == 0 else BLANCO

    # # ítem
    c = ws.cell(row=r, column=1)
    c.value     = i + 1
    c.font      = Font(size=10, bold=True, color=ROJO)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border_thin()

    # Descripción
    c = ws.cell(row=r, column=2)
    c.fill      = fill(bg)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border    = border_thin()
    c.font      = Font(size=10)

    # Precio proveedor
    c = ws.cell(row=r, column=3)
    c.fill           = fill(bg)
    c.alignment      = Alignment(horizontal="right", vertical="center")
    c.border         = border_thin()
    c.number_format  = '"$" #,##0'
    c.font           = Font(size=10)

    # % Rentabilidad — valor por defecto 65%
    c = ws.cell(row=r, column=4)
    c.value          = 0.65
    c.fill           = fill(ROJO_L)
    c.alignment      = Alignment(horizontal="center", vertical="center")
    c.border         = border_thin()
    c.number_format  = "0%"
    c.font           = Font(size=10, color=ROJO, bold=True)

    # Precio unitario = Precio proveedor / % rentabilidad
    col_c = get_column_letter(3)
    col_d = get_column_letter(4)
    c = ws.cell(row=r, column=5)
    c.value          = f'=IF(AND({col_c}{r}<>"",{col_d}{r}<>0),ROUND({col_c}{r}/{col_d}{r},0),"")'
    c.fill           = fill("E8F5E9")   # verde claro → precio de venta
    c.alignment      = Alignment(horizontal="right", vertical="center")
    c.border         = border_thin()
    c.number_format  = '"$" #,##0'
    c.font           = Font(size=10, bold=True, color="1B5E20")  # verde oscuro

    ws.row_dimensions[r].height = 22

# ── FILA VACÍA SEPARADORA ─────────────────────────────────────────────────
r_sep = 6 + FILAS
ws.row_dimensions[r_sep].height = 8

# ── NOTA AL PIE ───────────────────────────────────────────────────────────
r_nota = r_sep + 1
ws.merge_cells(f"A{r_nota}:E{r_nota}")
c = ws[f"A{r_nota}"]
c.value     = "FORMULA: Precio Unitario = Precio Proveedor / % Rentabilidad   |   Cambia el % en columna D para ajustar el margen"
c.font      = Font(size=9, italic=True, color="555555")
c.fill      = fill("FFFBEA")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r_nota].height = 18

# ── PROTEGER COLUMNAS DE FÓRMULAS (opcional: solo lectura E) ──────────────
# Dejamos todo editable para que sea flexible

# ── GUARDAR ───────────────────────────────────────────────────────────────
import winreg
try:
    k = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                       r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
    escritorio = winreg.QueryValueEx(k, "Desktop")[0]
except Exception:
    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
ruta = os.path.join(escritorio, "CALCULADORA_PRECIOS_MINZOE.xlsx")
wb.save(ruta)
print(f"OK - Archivo guardado en: {ruta}")
