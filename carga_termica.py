#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calculadora de Carga Termica HVAC - Construcciones Minzoe SAS
Basado en NTC 5183 y ASHRAE Fundamentals
"""
import tkinter as tk
from tkinter import ttk, messagebox
import math, os

# ═══════════════════════════════════════════════════════════════════
# DATOS NTC 5183 / ASHRAE
# ═══════════════════════════════════════════════════════════════════

# Temperaturas de diseno por ciudad colombiana (TDB grados C, TWB grados C)
CIUDADES = {
    "Barranquilla":  (34, 26), "Bogota":       (19, 13),
    "Bucaramanga":   (33, 22), "Cali":         (33, 22),
    "Cartagena":     (33, 27), "Cucuta":       (38, 25),
    "Ibague":        (32, 22), "Manizales":    (24, 17),
    "Medellin":      (29, 20), "Monteria":     (36, 27),
    "Neiva":         (35, 24), "Pasto":        (21, 14),
    "Pereira":       (28, 20), "Riohacha":     (37, 27),
    "Santa Marta":   (34, 26), "Sincelejo":    (35, 26),
    "Tunja":         (18, 12), "Valledupar":   (38, 26),
    "Villavicencio": (35, 25), "Otra ciudad":  (32, 23),
}

# Tipos de espacio: (BTU sens/persona, BTU lat/persona, CFM/persona, CFM/ft2)
# Fuente: ASHRAE Fundamentals + ASHRAE 62.1
TIPOS_ESPACIO = {
    "Oficina":              (250, 200, 10, 0.12),
    "Sala de reuniones":    (250, 200, 10, 0.12),
    "Local comercial":      (300, 250,  8, 0.12),
    "Restaurante":          (350, 400, 18, 0.18),
    "Sala de espera":       (250, 350, 10, 0.12),
    "Sala de computo":      (250, 200, 20, 0.12),
    "Clinica / Hospital":   (250, 300, 15, 0.18),
    "Aula / Salon":         (250, 350, 10, 0.12),
    "Bodega / Industrial":  (400, 400, 10, 0.06),
}

# Materiales paredes - U-value W/(m2*K)  Fuente: NTC 5183 / ASHRAE
MAT_PARED = {
    "Bloque concreto 15cm":          2.0,
    "Bloque concreto 20cm":          1.8,
    "Ladrillo comun 15cm":           2.2,
    "Concreto macizo 20cm":          3.5,
    "Panel metalico + aislamiento":  0.6,
    "Drywall doble + aislamiento":   0.8,
    "Ladrillo + estuco":             2.0,
}

# Materiales techo - U-value W/(m2*K)
MAT_TECHO = {
    "Losa concreto sin aislam.":       3.5,
    "Losa concreto + aislam.":         0.9,
    "Cubierta metalica sin aislam.":   6.0,
    "Cubierta metalica + aislam.":     0.8,
    "Teja fibrocemento sin aislam.":   5.5,
    "Teja fibrocemento + aislam.":     1.0,
    "Losa + cielo raso + aislam.":     0.7,
}

# Tipos de vidrio: (SHGC, U W/m2K)  Fuente: ASHRAE
VIDRIOS = {
    "Vidrio sencillo claro":          (0.86, 5.8),
    "Vidrio sencillo bronce/gris":    (0.55, 5.8),
    "Vidrio reflectivo plateado":     (0.25, 5.8),
    "Vidrio doble claro":             (0.76, 2.8),
    "Vidrio doble reflectivo":        (0.35, 2.8),
    "Vidrio Low-E (bajo emisivo)":    (0.27, 1.6),
}

# Irradiancia solar W/m2 por orientacion (Colombia, latitud 4-11 grados N)
ORIENT_SOLAR = {"Norte": 100, "Sur": 150, "Este": 500, "Oeste": 500}

# Tamanos estandar de equipos en BTU
TALLAS = [9000, 12000, 18000, 24000, 30000, 36000, 48000, 60000, 120000, 240000]

# ═══════════════════════════════════════════════════════════════════
# CALCULO
# ═══════════════════════════════════════════════════════════════════

def calcular(d):
    t_out = d["t_out"];  t_in = d["t_in"]
    dt    = max(t_out - t_in, 0)       # diferencia C
    dt_f  = dt * 9 / 5                 # diferencia F
    area  = d["area"]
    area_ft2 = area * 10.764

    # 1. Paredes
    u_p = MAT_PARED.get(d["mat_pared"], 2.0)
    q_pared = u_p * d["a_pared"] * dt * 3.412

    # 2. Techo
    u_t = MAT_TECHO.get(d["mat_techo"], 3.5)
    q_techo = u_t * area * dt * 3.412

    # 3. Ventanas - conduccion
    shgc, u_v = VIDRIOS.get(d["tipo_vidrio"], (0.86, 5.8))
    a_vent = d["a_ventanas"]
    q_vid_cond = u_v * a_vent * dt * 3.412

    # 4. Ventanas - solar
    g_solar = ORIENT_SOLAR.get(d["orientacion"], 300)
    q_solar = shgc * g_solar * a_vent * 3.412

    # 5. Personas
    tipo_esp = d["tipo_espacio"]
    sens_pp, lat_pp, cfm_pp, cfm_m2 = TIPOS_ESPACIO.get(tipo_esp, (250, 200, 10, 0.12))
    n = d["personas"]
    q_pers_s = n * sens_pp
    q_pers_l = n * lat_pp

    # 6. Iluminacion  (W/m2 x area x factor 3.412 BTU/W)
    q_ilum = d["w_m2"] * area * 3.412

    # 7. Equipos
    q_equip = (d["n_pc"]    * 400  +
               d["n_lap"]   * 200  +
               d["n_imp_p"] * 600  +
               d["n_imp_g"] * 1200 +
               d["n_srv"]   * 2000 +
               d["n_tv"]    * 400  +
               d["n_cop"]   * 1500 +
               d["otros_w"] * 3.412)

    # 8. Ventilacion (ASHRAE 62.1)
    cfm_vent = n * cfm_pp + area_ft2 * cfm_m2
    q_vent_s = cfm_vent * 1.1 * dt_f
    dw       = d.get("delta_w", 50)          # gr/lb diferencia humedad
    q_vent_l = cfm_vent * 0.68 * dw

    # 9. Totales
    q_sens = (q_pared + q_techo + q_vid_cond + q_solar +
              q_pers_s + q_ilum + q_equip + q_vent_s)
    q_lat  = q_pers_l + q_vent_l
    q_sub  = q_sens + q_lat
    fs     = d["factor_seg"] / 100
    q_total = q_sub * (1 + fs)

    # 10. Resultados
    tons  = q_total / 12000
    cfm_s = tons * 400
    m3h   = cfm_s * 1.699
    shr   = q_sens / max(q_sub, 1)

    # 11. Seleccion equipo
    eq_btu = next((t for t in TALLAS if t >= q_total), TALLAS[-1])
    if eq_btu <= 18000:
        tipo_eq = f"Minisplit {eq_btu // 1000}k BTU"
    elif eq_btu <= 60000:
        tipo_eq = f"Minisplit/Cassette {eq_btu // 1000}k BTU"
    else:
        nu = math.ceil(q_total / 60000)
        tipo_eq = f"{nu} unidades de 60,000 BTU"

    return {
        "q_pared": q_pared, "q_techo": q_techo,
        "q_vid_cond": q_vid_cond, "q_solar": q_solar,
        "q_pers_s": q_pers_s, "q_pers_l": q_pers_l,
        "q_ilum": q_ilum, "q_equip": q_equip,
        "q_vent_s": q_vent_s, "q_vent_l": q_vent_l,
        "q_sens": q_sens, "q_lat": q_lat,
        "q_total": q_total, "tons": tons,
        "cfm": cfm_s, "m3h": m3h, "shr": shr,
        "eq_btu": eq_btu, "tipo_eq": tipo_eq,
        "cfm_vent": cfm_vent,
    }

# ═══════════════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════════════
ROJO   = "#DC2626"
BLANCO = "#FFFFFF"
GRIS   = "#F8F8F8"
VERDE  = "#166534"

def lbl(p, txt, bold=False, color="#111", size=9):
    return tk.Label(p, text=txt, font=("Segoe UI", size, "bold" if bold else "normal"),
                    fg=color, bg=p.cget("bg") if hasattr(p, 'cget') else GRIS)

def seccion(p, txt, row):
    tk.Label(p, text=txt, font=("Segoe UI", 9, "bold"), fg=BLANCO,
             bg=ROJO, anchor="w", padx=6
             ).grid(row=row, column=0, columnspan=2, sticky="ew",
                    pady=(10, 2), ipady=2)

def fila(p, label, widget, row, nota=None):
    tk.Label(p, text=label, font=("Segoe UI", 9), fg="#333",
             bg=p.cget("bg"), anchor="w"
             ).grid(row=row, column=0, sticky="w", pady=2, padx=(4, 0))
    widget.grid(row=row, column=1, sticky="w", padx=6, pady=2)
    if nota:
        tk.Label(p, text=nota, font=("Segoe UI", 7), fg="#999",
                 bg=p.cget("bg")).grid(row=row, column=2, sticky="w")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Carga Termica HVAC - Minzoe")
        self.configure(bg=GRIS)
        self.resizable(True, True)
        self._resultados = None
        self._datos      = None
        self._build_ui()
        w, h = 980, 700
        self.geometry(f"{w}x{h}+{(self.winfo_screenwidth()-w)//2}+{(self.winfo_screenheight()-h)//2}")

    # ── UI ────────────────────────────────────────────────────────────
    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=ROJO, height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="CONSTRUCCIONES MINZOE SAS",
                 font=("Segoe UI", 13, "bold"), fg=BLANCO, bg=ROJO
                 ).pack(side="left", padx=16, pady=10)
        tk.Label(hdr, text="Calculadora Carga Termica HVAC  |  NTC 5183 + ASHRAE",
                 font=("Segoe UI", 9), fg="#FFCDD2", bg=ROJO
                 ).pack(side="right", padx=14)

        # Contenedor principal
        main = tk.Frame(self, bg=GRIS)
        main.pack(fill="both", expand=True, padx=8, pady=6)
        main.columnconfigure(0, weight=3)
        main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        # Panel izquierdo (inputs)
        left = tk.Frame(main, bg=GRIS)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 4))

        nb = ttk.Notebook(left)
        nb.pack(fill="both", expand=True)

        t1 = tk.Frame(nb, bg=GRIS, padx=8, pady=6)
        t2 = tk.Frame(nb, bg=GRIS, padx=8, pady=6)
        t3 = tk.Frame(nb, bg=GRIS, padx=8, pady=6)
        nb.add(t1, text="  Espacio  ")
        nb.add(t2, text="  Envolvente  ")
        nb.add(t3, text="  Cargas Internas  ")
        self._tab_espacio(t1)
        self._tab_envolvente(t2)
        self._tab_cargas(t3)

        # Botones
        bf = tk.Frame(left, bg=ROJO)
        bf.pack(fill="x", pady=(4, 0))
        tk.Button(bf, text="  CALCULAR  ",
                  font=("Segoe UI", 11, "bold"),
                  bg=ROJO, fg=BLANCO, activebackground="#B91C1C",
                  relief="flat", cursor="hand2", padx=12, pady=6,
                  command=self.calcular
                  ).pack(side="left", padx=8, pady=4)
        tk.Button(bf, text="  Exportar Excel  ",
                  font=("Segoe UI", 9),
                  bg="#7F1D1D", fg=BLANCO, activebackground="#991B1B",
                  relief="flat", cursor="hand2", padx=8, pady=6,
                  command=self.exportar
                  ).pack(side="right", padx=8, pady=4)
        tk.Button(bf, text="  Limpiar  ",
                  font=("Segoe UI", 9),
                  bg="#7F1D1D", fg=BLANCO, activebackground="#991B1B",
                  relief="flat", cursor="hand2", padx=8, pady=6,
                  command=self.limpiar
                  ).pack(side="right", padx=0, pady=4)

        # Panel derecho (resultados)
        right = tk.Frame(main, bg=GRIS)
        right.grid(row=0, column=1, sticky="nsew")
        self._build_resultados(right)

    # ── Tab 1: Espacio ────────────────────────────────────────────────
    def _tab_espacio(self, p):
        p.columnconfigure(1, weight=1)

        seccion(p, "UBICACION Y CONDICIONES", 0)
        self.v_ciudad = tk.StringVar(value="Monteria")
        cb = ttk.Combobox(p, textvariable=self.v_ciudad,
                           values=list(CIUDADES.keys()), width=22, state="readonly")
        cb.bind("<<ComboboxSelected>>", self._on_ciudad)
        fila(p, "Ciudad:", cb, 1)

        self.v_t_out = tk.StringVar(value="36")
        fila(p, "Temp. exterior diseno (C):", ttk.Entry(p, textvariable=self.v_t_out, width=8), 2,
             "Auto segun ciudad")

        self.v_t_in = tk.StringVar(value="24")
        fila(p, "Temp. interior (C):", ttk.Entry(p, textvariable=self.v_t_in, width=8), 3,
             "Recomendado: 22-24 C")

        self.v_dw = tk.StringVar(value="50")
        fila(p, "Delta humedad (gr/lb):", ttk.Entry(p, textvariable=self.v_dw, width=8), 4,
             "Costa: 60-80 | Interior: 40-60")

        seccion(p, "DATOS DEL ESPACIO", 5)
        self.v_tipo = tk.StringVar(value="Oficina")
        fila(p, "Tipo de espacio:", ttk.Combobox(p, textvariable=self.v_tipo,
             values=list(TIPOS_ESPACIO.keys()), width=22, state="readonly"), 6)

        self.v_area = tk.StringVar(value="24")
        fila(p, "Area total m2:", ttk.Entry(p, textvariable=self.v_area, width=8), 7)

        self.v_altura = tk.StringVar(value="3")
        fila(p, "Altura m:", ttk.Entry(p, textvariable=self.v_altura, width=8), 8)

        seccion(p, "FACTOR DE SEGURIDAD", 9)
        self.v_fs = tk.StringVar(value="10")
        fila(p, "Factor de seguridad %:", ttk.Entry(p, textvariable=self.v_fs, width=8), 10,
             "NTC 5183 recomienda 10-15%")

    def _on_ciudad(self, _=None):
        c = self.v_ciudad.get()
        if c in CIUDADES:
            self.v_t_out.set(str(CIUDADES[c][0]))

    # ── Tab 2: Envolvente ─────────────────────────────────────────────
    def _tab_envolvente(self, p):
        p.columnconfigure(1, weight=1)

        seccion(p, "PAREDES", 0)
        self.v_mat_p = tk.StringVar(value=list(MAT_PARED.keys())[0])
        fila(p, "Material:", ttk.Combobox(p, textvariable=self.v_mat_p,
             values=list(MAT_PARED.keys()), width=28, state="readonly"), 1)
        self.v_a_pared = tk.StringVar(value="55")
        fila(p, "Area paredes m2 (sin ventanas):", ttk.Entry(p, textvariable=self.v_a_pared, width=8), 2,
             "Perimetro x altura - ventanas")

        seccion(p, "TECHO / CUBIERTA", 3)
        self.v_mat_t = tk.StringVar(value=list(MAT_TECHO.keys())[0])
        fila(p, "Material:", ttk.Combobox(p, textvariable=self.v_mat_t,
             values=list(MAT_TECHO.keys()), width=28, state="readonly"), 4)
        tk.Label(p, text="  Area = misma que area del espacio (se usa automaticamente)",
                 font=("Segoe UI", 7), fg="#999", bg=GRIS
                 ).grid(row=5, column=0, columnspan=3, sticky="w")

        seccion(p, "VENTANAS", 6)
        self.v_vid = tk.StringVar(value=list(VIDRIOS.keys())[0])
        fila(p, "Tipo de vidrio:", ttk.Combobox(p, textvariable=self.v_vid,
             values=list(VIDRIOS.keys()), width=28, state="readonly"), 7)
        self.v_a_vent = tk.StringVar(value="0")
        fila(p, "Area total ventanas m2:", ttk.Entry(p, textvariable=self.v_a_vent, width=8), 8)
        self.v_orient = tk.StringVar(value="Oeste")
        fila(p, "Orientacion principal:", ttk.Combobox(p, textvariable=self.v_orient,
             values=list(ORIENT_SOLAR.keys()), width=14, state="readonly"), 9,
             "Oeste/Este = mayor ganancia solar")

        tk.Label(p, text=(
            "\n  Valores U (W/m2K) segun NTC 5183 / ASHRAE:\n"
            "  Bloque 15cm: 2.0  |  Losa sin aislam.: 3.5  |  Cubierta metal sin aislam.: 6.0\n"
            "  SHGC vidrio sencillo: 0.86  |  Reflectivo: 0.25  |  Low-E: 0.27"
        ), font=("Segoe UI", 7), fg="#777", bg=GRIS, justify="left"
        ).grid(row=10, column=0, columnspan=3, sticky="w", pady=(10, 0))

    # ── Tab 3: Cargas internas ────────────────────────────────────────
    def _tab_cargas(self, p):
        p.columnconfigure(1, weight=1)

        seccion(p, "PERSONAS", 0)
        self.v_pers = tk.StringVar(value="4")
        fila(p, "Numero de personas:", ttk.Entry(p, textvariable=self.v_pers, width=8), 1,
             "Calor segun tipo de espacio (ASHRAE)")

        seccion(p, "ILUMINACION", 2)
        self.v_wm2 = tk.StringVar(value="10")
        fila(p, "Carga iluminacion W/m2:", ttk.Entry(p, textvariable=self.v_wm2, width=8), 3,
             "LED: 8-12 | Fluorescente: 15-20 | Halogeno: 25-30")

        seccion(p, "EQUIPOS (unidades)", 4)
        self.v_pc    = tk.StringVar(value="4");  fila(p, "Computadores escritorio (400 BTU):",  ttk.Entry(p, textvariable=self.v_pc,    width=6), 5)
        self.v_lap   = tk.StringVar(value="0");  fila(p, "Laptops / Notebooks (200 BTU):",      ttk.Entry(p, textvariable=self.v_lap,   width=6), 6)
        self.v_imp_p = tk.StringVar(value="0");  fila(p, "Impresoras pequenas (600 BTU):",      ttk.Entry(p, textvariable=self.v_imp_p, width=6), 7)
        self.v_imp_g = tk.StringVar(value="1");  fila(p, "Impresoras industriales (1200 BTU):", ttk.Entry(p, textvariable=self.v_imp_g, width=6), 8)
        self.v_srv   = tk.StringVar(value="0");  fila(p, "Servidores / UPS (2000 BTU):",        ttk.Entry(p, textvariable=self.v_srv,   width=6), 9)
        self.v_tv    = tk.StringVar(value="0");  fila(p, "TV / Monitores grandes (400 BTU):",   ttk.Entry(p, textvariable=self.v_tv,    width=6), 10)
        self.v_cop   = tk.StringVar(value="0");  fila(p, "Fotocopiadoras (1500 BTU):",          ttk.Entry(p, textvariable=self.v_cop,   width=6), 11)
        self.v_otros = tk.StringVar(value="0");  fila(p, "Otros equipos (Watts totales):",      ttk.Entry(p, textvariable=self.v_otros, width=6), 12,
                                                      "1 W = 3.412 BTU/hr")

    # ── Panel resultados ──────────────────────────────────────────────
    def _build_resultados(self, p):
        tk.Label(p, text="RESULTADOS", font=("Segoe UI", 10, "bold"),
                 fg=BLANCO, bg=ROJO, anchor="w", padx=8
                 ).pack(fill="x", ipady=4)

        canvas = tk.Canvas(p, bg=GRIS, highlightthickness=0)
        sb = ttk.Scrollbar(p, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=GRIS)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _resize(e):
            canvas.itemconfig(win_id, width=e.width)
        canvas.bind("<Configure>", _resize)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        self.res = {}
        filas_res = [
            ("sep1",     "--- SENSIBLES ---"),
            ("q_pared",  "Paredes"),
            ("q_techo",  "Techo/Cubierta"),
            ("q_vid",    "Ventanas conduccion"),
            ("q_solar",  "Ventanas solar"),
            ("q_pers_s", "Personas sensible"),
            ("q_ilum",   "Iluminacion"),
            ("q_equip",  "Equipos"),
            ("q_vent_s", "Ventilacion sensible"),
            ("q_sens",   "TOTAL SENSIBLE"),
            ("sep2",     "--- LATENTES ---"),
            ("q_pers_l", "Personas latente"),
            ("q_vent_l", "Ventilacion latente"),
            ("q_lat",    "TOTAL LATENTE"),
            ("sep3",     "--- TOTALES ---"),
            ("shr",      "SHR (ratio sensible)"),
            ("q_total",  "CARGA TOTAL"),
            ("tons",     "Toneladas"),
            ("cfm",      "CFM requeridos"),
            ("m3h",      "m3/h requeridos"),
            ("cfm_v",    "CFM ventilacion minima"),
            ("sep4",     "--- EQUIPO ---"),
            ("eq_btu",   "BTU del equipo"),
            ("tipo_eq",  "Tipo de equipo"),
        ]
        TOTAL_KEYS = {"q_sens", "q_lat", "q_total", "tons", "tipo_eq"}
        for i, (key, label) in enumerate(filas_res):
            if key.startswith("sep"):
                tk.Label(inner, text=label, font=("Courier New", 7),
                         fg="#AAAAAA", bg=GRIS, anchor="w"
                         ).grid(row=i, column=0, columnspan=2, sticky="ew",
                                padx=6, pady=(8, 0))
            else:
                bold  = key in TOTAL_KEYS
                color = ROJO if key == "q_total" else (VERDE if bold else "#222")
                tk.Label(inner, text=label + ":", font=("Segoe UI", 9),
                         fg="#555", bg=GRIS, anchor="w", width=22
                         ).grid(row=i, column=0, sticky="w", padx=(8, 0), pady=1)
                vl = tk.Label(inner, text="—", font=("Segoe UI", 9, "bold" if bold else "normal"),
                              fg=color, bg=GRIS, anchor="e")
                vl.grid(row=i, column=1, sticky="e", padx=(0, 10), pady=1)
                self.res[key] = vl
        inner.columnconfigure(1, weight=1)

    # ── Helpers ───────────────────────────────────────────────────────
    def _f(self, var, default=0.0):
        try:   return float(str(var.get()).replace(",", "."))
        except: return default

    def _i(self, var, default=0):
        try:   return int(float(str(var.get()).replace(",", ".")))
        except: return default

    # ── Calcular ──────────────────────────────────────────────────────
    def calcular(self):
        datos = {
            "t_out":      self._f(self.v_t_out, 32),
            "t_in":       self._f(self.v_t_in,  24),
            "tipo_espacio": self.v_tipo.get(),
            "area":       self._f(self.v_area,  24),
            "mat_pared":  self.v_mat_p.get(),
            "a_pared":    self._f(self.v_a_pared, 55),
            "mat_techo":  self.v_mat_t.get(),
            "tipo_vidrio": self.v_vid.get(),
            "a_ventanas": self._f(self.v_a_vent, 0),
            "orientacion": self.v_orient.get(),
            "personas":   self._i(self.v_pers, 0),
            "w_m2":       self._f(self.v_wm2, 10),
            "n_pc":       self._i(self.v_pc,    0),
            "n_lap":      self._i(self.v_lap,   0),
            "n_imp_p":    self._i(self.v_imp_p, 0),
            "n_imp_g":    self._i(self.v_imp_g, 0),
            "n_srv":      self._i(self.v_srv,   0),
            "n_tv":       self._i(self.v_tv,    0),
            "n_cop":      self._i(self.v_cop,   0),
            "otros_w":    self._f(self.v_otros, 0),
            "factor_seg": self._f(self.v_fs,    10),
            "delta_w":    self._f(self.v_dw,    50),
        }
        r = calcular(datos)
        self._resultados = r
        self._datos      = datos

        def btu(v): return f"{v:,.0f} BTU/hr"

        self.res["q_pared"].config(text=btu(r["q_pared"]))
        self.res["q_techo"].config(text=btu(r["q_techo"]))
        self.res["q_vid"].config(text=btu(r["q_vid_cond"]))
        self.res["q_solar"].config(text=btu(r["q_solar"]))
        self.res["q_pers_s"].config(text=btu(r["q_pers_s"]))
        self.res["q_ilum"].config(text=btu(r["q_ilum"]))
        self.res["q_equip"].config(text=btu(r["q_equip"]))
        self.res["q_vent_s"].config(text=btu(r["q_vent_s"]))
        self.res["q_sens"].config(text=btu(r["q_sens"]))
        self.res["q_pers_l"].config(text=btu(r["q_pers_l"]))
        self.res["q_vent_l"].config(text=btu(r["q_vent_l"]))
        self.res["q_lat"].config(text=btu(r["q_lat"]))
        self.res["shr"].config(text=f"{r['shr']:.3f}")
        self.res["q_total"].config(text=btu(r["q_total"]))
        self.res["tons"].config(text=f"{r['tons']:.3f} Ton")
        self.res["cfm"].config(text=f"{r['cfm']:,.1f} CFM")
        self.res["m3h"].config(text=f"{r['m3h']:,.1f} m3/h")
        self.res["cfm_v"].config(text=f"{r['cfm_vent']:,.1f} CFM")
        self.res["eq_btu"].config(text=f"{r['eq_btu']:,} BTU")
        self.res["tipo_eq"].config(text=r["tipo_eq"])

    def limpiar(self):
        for v, d in [(self.v_area,"24"),(self.v_altura,"3"),(self.v_pers,"4"),
                     (self.v_wm2,"10"),(self.v_a_pared,"55"),(self.v_a_vent,"0"),
                     (self.v_pc,"0"),(self.v_lap,"0"),(self.v_imp_p,"0"),
                     (self.v_imp_g,"0"),(self.v_srv,"0"),(self.v_tv,"0"),
                     (self.v_cop,"0"),(self.v_otros,"0")]:
            v.set(d)
        for lbl_w in self.res.values():
            lbl_w.config(text="—", fg=VERDE)
        self.res["q_total"].config(fg=ROJO)
        self._resultados = None

    # ── Exportar Excel ────────────────────────────────────────────────
    def exportar(self):
        if not self._resultados:
            messagebox.showwarning("Aviso", "Primero haz clic en CALCULAR.")
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Error", "Instala openpyxl:\n pip install openpyxl")
            return

        r = self._resultados
        d = self._datos
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Carga Termica"
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 22

        fill_r = PatternFill("solid", fgColor="DC2626")
        fill_g = PatternFill("solid", fgColor="F8F8F8")
        fill_v = PatternFill("solid", fgColor="DCFCE7")
        thin   = lambda: Border(*[Side(style="thin", color="DDDDDD")]*0,
                                left=Side(style="thin",color="DDDDDD"),
                                right=Side(style="thin",color="DDDDDD"),
                                top=Side(style="thin",color="DDDDDD"),
                                bottom=Side(style="thin",color="DDDDDD"))

        def hdr(r_num, txt):
            ws.merge_cells(f"A{r_num}:B{r_num}")
            c = ws[f"A{r_num}"]
            c.value = txt
            c.font  = Font(bold=True, color="FFFFFF", size=10)
            c.fill  = fill_r
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r_num].height = 18

        def dato(r_num, lab, val, bold=False, fill_=None):
            ca, cb = ws[f"A{r_num}"], ws[f"B{r_num}"]
            ca.value = lab; cb.value = val
            for c in (ca, cb):
                c.border = thin()
                c.font   = Font(bold=bold, size=9)
                if fill_: c.fill = fill_
            cb.alignment = Alignment(horizontal="right")
            ws.row_dimensions[r_num].height = 16

        # Titulo
        ws.merge_cells("A1:B1")
        c = ws["A1"]
        c.value = "CONSTRUCCIONES MINZOE SAS - Calculo Carga Termica HVAC"
        c.font  = Font(bold=True, size=12, color="FFFFFF")
        c.fill  = fill_r
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        ws.merge_cells("A2:B2")
        ws["A2"].value = "Basado en NTC 5183 y ASHRAE Fundamentals"
        ws["A2"].font  = Font(size=9, color="DC2626", italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        n = 4
        hdr(n, "DATOS DE ENTRADA"); n+=1
        dato(n,"Ciudad",                    self.v_ciudad.get()); n+=1
        dato(n,"Temperatura exterior (C)",  d["t_out"]); n+=1
        dato(n,"Temperatura interior (C)",  d["t_in"]); n+=1
        dato(n,"Tipo de espacio",           d["tipo_espacio"]); n+=1
        dato(n,"Area m2",                   d["area"]); n+=1
        dato(n,"Material paredes",          d["mat_pared"]); n+=1
        dato(n,"Area paredes m2",           d["a_pared"]); n+=1
        dato(n,"Material techo",            d["mat_techo"]); n+=1
        dato(n,"Tipo de vidrio",            d["tipo_vidrio"]); n+=1
        dato(n,"Area ventanas m2",          d["a_ventanas"]); n+=1
        dato(n,"Orientacion",               d["orientacion"]); n+=1
        dato(n,"Personas",                  d["personas"]); n+=1
        dato(n,"Iluminacion W/m2",          d["w_m2"]); n+=1
        dato(n,"Factor de seguridad %",     d["factor_seg"]); n+=1

        n+=1; hdr(n,"CARGAS SENSIBLES (BTU/hr)"); n+=1
        dato(n,"Paredes",                   f"{r['q_pared']:,.0f}"); n+=1
        dato(n,"Techo/Cubierta",            f"{r['q_techo']:,.0f}"); n+=1
        dato(n,"Ventanas (conduccion)",     f"{r['q_vid_cond']:,.0f}"); n+=1
        dato(n,"Ventanas (solar)",          f"{r['q_solar']:,.0f}"); n+=1
        dato(n,"Personas",                  f"{r['q_pers_s']:,.0f}"); n+=1
        dato(n,"Iluminacion",               f"{r['q_ilum']:,.0f}"); n+=1
        dato(n,"Equipos",                   f"{r['q_equip']:,.0f}"); n+=1
        dato(n,"Ventilacion (sensible)",    f"{r['q_vent_s']:,.0f}"); n+=1
        dato(n,"TOTAL SENSIBLE",            f"{r['q_sens']:,.0f}", bold=True, fill_=fill_g); n+=1

        n+=1; hdr(n,"CARGAS LATENTES (BTU/hr)"); n+=1
        dato(n,"Personas (latente)",        f"{r['q_pers_l']:,.0f}"); n+=1
        dato(n,"Ventilacion (latente)",     f"{r['q_vent_l']:,.0f}"); n+=1
        dato(n,"TOTAL LATENTE",             f"{r['q_lat']:,.0f}", bold=True, fill_=fill_g); n+=1

        n+=1; hdr(n,"RESULTADOS FINALES"); n+=1
        dato(n,"SHR (Sensible Heat Ratio)", f"{r['shr']:.3f}"); n+=1
        dato(n,"CARGA TOTAL (BTU/hr)",      f"{r['q_total']:,.0f}", bold=True, fill_=fill_v); n+=1
        dato(n,"Toneladas de refrig.",       f"{r['tons']:.3f} Ton", bold=True, fill_=fill_v); n+=1
        dato(n,"CFM requeridos",            f"{r['cfm']:,.1f}"); n+=1
        dato(n,"m3/h requeridos",           f"{r['m3h']:,.1f}"); n+=1
        dato(n,"CFM ventilacion (ASHRAE 62.1)", f"{r['cfm_vent']:,.1f}"); n+=1
        dato(n,"Equipo recomendado (BTU)",  f"{r['eq_btu']:,}"); n+=1
        dato(n,"Tipo de equipo",            r["tipo_eq"], bold=True, fill_=fill_v); n+=1

        try:
            import winreg
            k = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                               r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
            desk = winreg.QueryValueEx(k, "Desktop")[0]
        except Exception:
            desk = os.path.expanduser("~")

        ciudad_safe = self.v_ciudad.get().replace(" ", "_").replace("/", "-")
        ruta = os.path.join(desk, f"CargaTermica_{ciudad_safe}.xlsx")
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{ruta}")


if __name__ == "__main__":
    App().mainloop()
